from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Q, Avg, Sum
from django.utils import timezone
from datetime import datetime, timedelta
import json
import csv
import openpyxl
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from io import BytesIO

from .models import Project, TestReport, ThicknessData, WeldData, Company, Equipment
from .permissions import staff_required

@staff_required
def report_dashboard(request):
    """Dashboard báo cáo tổng quan"""
    # Thống kê tổng quan
    total_projects = Project.objects.count()
    active_projects = Project.objects.filter(status='active').count()
    completed_projects = Project.objects.filter(status='completed').count()
    total_reports = TestReport.objects.count()
    
    # Thống kê theo tháng
    current_month = timezone.now().month
    current_year = timezone.now().year
    monthly_projects = Project.objects.filter(
        created_at__month=current_month,
        created_at__year=current_year
    ).count()
    
    # Thống kê theo công ty
    company_stats = Company.objects.annotate(
        project_count=Count('project')
    ).order_by('-project_count')[:5]
    
    # Thống kê thiết bị
    equipment_usage = Equipment.objects.annotate(
        usage_count=Count('project')
    ).order_by('-usage_count')[:5]
    
    context = {
        'total_projects': total_projects,
        'active_projects': active_projects,
        'completed_projects': completed_projects,
        'total_reports': total_reports,
        'monthly_projects': monthly_projects,
        'company_stats': company_stats,
        'equipment_usage': equipment_usage,
    }
    
    return render(request, 'reports/dashboard.html', context)

@staff_required
def project_report(request):
    """Báo cáo dự án"""
    # Lọc dữ liệu
    projects = Project.objects.all()
    
    # Áp dụng bộ lọc
    company_id = request.GET.get('company')
    status = request.GET.get('status')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if company_id:
        projects = projects.filter(company_id=company_id)
    if status:
        projects = projects.filter(status=status)
    if start_date:
        projects = projects.filter(start_date__gte=start_date)
    if end_date:
        projects = projects.filter(end_date__lte=end_date)
    
    # Xuất báo cáo
    export_format = request.GET.get('export')
    if export_format:
        return export_project_report(projects, export_format)
    
    context = {
        'projects': projects,
        'companies': Company.objects.all(),
        'status_choices': Project.STATUS_CHOICES,
    }
    
    return render(request, 'reports/project_report.html', context)

@staff_required
def thickness_report(request):
    """Báo cáo dữ liệu độ dày"""
    thickness_data = ThicknessData.objects.select_related('project', 'inspector', 'equipment')
    
    # Áp dụng bộ lọc
    project_id = request.GET.get('project')
    inspector_id = request.GET.get('inspector')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    if project_id:
        thickness_data = thickness_data.filter(project_id=project_id)
    if inspector_id:
        thickness_data = thickness_data.filter(inspector_id=inspector_id)
    if date_from:
        thickness_data = thickness_data.filter(measurement_date__gte=date_from)
    if date_to:
        thickness_data = thickness_data.filter(measurement_date__lte=date_to)
    
    # Thống kê
    total_measurements = thickness_data.count()
    acceptable_count = thickness_data.filter(
        thickness_value__gte=models.F('minimum_thickness')
    ).count()
    
    acceptance_rate = (acceptable_count / total_measurements * 100) if total_measurements > 0 else 0
    
    # Xuất báo cáo
    export_format = request.GET.get('export')
    if export_format:
        return export_thickness_report(thickness_data, export_format)
    
    context = {
        'thickness_data': thickness_data,
        'total_measurements': total_measurements,
        'acceptable_count': acceptable_count,
        'acceptance_rate': round(acceptance_rate, 2),
        'projects': Project.objects.all(),
    }
    
    return render(request, 'reports/thickness_report.html', context)

def export_project_report(projects, format_type):
    """Xuất báo cáo dự án theo định dạng"""
    if format_type == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="project_report.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Mã dự án', 'Tên dự án', 'Công ty', 'Trạng thái', 'Ngày bắt đầu', 'Ngày kết thúc', 'Giá trị hợp đồng'])
        
        for project in projects:
            writer.writerow([
                project.code,
                project.name,
                project.company.name,
                project.get_status_display(),
                project.start_date,
                project.end_date,
                project.contract_value or 0
            ])
        
        return response
    
    elif format_type == 'excel':
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="project_report.xlsx"'
        
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'Báo cáo dự án'
        
        # Header
        headers = ['Mã dự án', 'Tên dự án', 'Công ty', 'Trạng thái', 'Ngày bắt đầu', 'Ngày kết thúc', 'Giá trị hợp đồng']
        for col, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col, value=header)
        
        # Data
        for row, project in enumerate(projects, 2):
            worksheet.cell(row=row, column=1, value=project.code)
            worksheet.cell(row=row, column=2, value=project.name)
            worksheet.cell(row=row, column=3, value=project.company.name)
            worksheet.cell(row=row, column=4, value=project.get_status_display())
            worksheet.cell(row=row, column=5, value=project.start_date)
            worksheet.cell(row=row, column=6, value=project.end_date)
            worksheet.cell(row=row, column=7, value=float(project.contract_value or 0))
        
        workbook.save(response)
        return response
    
    elif format_type == 'pdf':
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="project_report.pdf"'
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        
        # Tạo bảng dữ liệu
        data = [['Mã dự án', 'Tên dự án', 'Công ty', 'Trạng thái']]
        for project in projects:
            data.append([
                project.code,
                project.name[:30] + '...' if len(project.name) > 30 else project.name,
                project.company.name[:20] + '...' if len(project.company.name) > 20 else project.company.name,
                project.get_status_display()
            ])
        
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        doc.build([table])
        pdf = buffer.getvalue()
        buffer.close()
        response.write(pdf)
        
        return response

def export_thickness_report(thickness_data, format_type):
    """Xuất báo cáo độ dày theo định dạng"""
    # Tương tự như export_project_report nhưng cho dữ liệu độ dày
    pass